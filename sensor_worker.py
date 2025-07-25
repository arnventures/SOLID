"""
sensor_worker.py – schnell & robust
-----------------------------------
Ablauf pro Sensor

1. Warten auf *neuen* Sensor @1   (Serial noch nicht gesehen)
2. Serien-Nr. lesen  → sofort in Blatt „Import“ (Spalte E) schreiben
3. Adresse/Buzzer setzen, Reboot senden
4. 2 s Pause, OK zurückgeben
"""

from __future__ import annotations
import time
import openpyxl
from typing import List, Dict, Callable
from serial_manager import SerialManager


BOOT_WAIT_S = 2.0          # feste Boot-Pause nach Reboot


class SensorWorker:
    def __init__(
        self,
        rows: List[Dict],                           # Sensordaten (aus GAS)
        workbook: openpyxl.Workbook,               # gesamtes Workbook
        worksheet: openpyxl.worksheet.worksheet.Worksheet,  # GAS-Sheet
        ser: SerialManager,
        log: Callable[[str], None],
        stop_event,
        skip_event,
    ):
        self.rows        = rows
        self.wb          = workbook
        self.ws_gas      = worksheet               # bleibt zum Lesen erhalten
        self.ws_import   = workbook["Import"]      # Serien-Nr. hier speichern

        self.ser         = ser
        self.log         = log
        self.stop_event  = stop_event
        self.skip_event  = skip_event
        self.prev_serials: set[int] = set()

    # ------------------------------------------------------------------ #
    def run(self):
        self.log("Worker gestartet.")
        for item in self.rows:
            if not item["enabled"]:
                self.log(f"Sensor {item['row']-1} übersprungen (deselektiert).")
                continue
            if self.stop_event.is_set():
                break

            if not self._wait_for_addr1():
                break

            ok, serial = self._configure(item)
            item["status_cb"](ok)

            if ok and serial:
                self.prev_serials.add(serial)
                self.wb.save(self.wb.filename)

        self.log("Worker beendet.")

    # ------------------------------------------------------------------ #
    def _wait_for_addr1(self) -> bool:
        self.log("Warte auf *neues* Gerät @1 …")
        while not self.stop_event.is_set():
            if self.skip_event.is_set():
                self.skip_event.clear()
                self.log("Warten abgebrochen (Skip).")
                return False
            try:
                res = self.ser.read_holding(3, unit=1)
                if not res.isError():
                    serial = res.registers[0]
                    if serial not in self.prev_serials:
                        self.log(f"Gerät @1 gefunden (SN={serial}).")
                        return True
            except Exception as e:
                self.log(f"Modbus-Fehler (wait): {e}")
            time.sleep(1)
        return False

    # ------------------------------------------------------------------ #
    def _configure(self, item: Dict) -> tuple[bool, int | None]:
        """
        • Serial sofort in „Import“ schreiben + GUI-Callback
        • Adresse/Buzzer setzen, Reboot senden
        • 2 s warten, OK zurück
        """
        row, new_addr, disable_bz = (
            item["row"], item["new_addr"], item["buzzer"]
        )

        if self.skip_event.is_set():
            self.skip_event.clear()
            return False, None

        self.log(f"Config {row-1} → Adresse {new_addr}")
        try:
            # 1) Serien-Nr. lesen (@1)
            res = self.ser.read_holding(3, unit=1)
            if res.isError():
                raise RuntimeError(res)
            serial = res.registers[0]

            # 1a) SOFORT in Import!E schreiben
            import_row = max(1, row - 1)
            if import_row > self.ws_import.max_row:
                while self.ws_import.max_row < import_row:
                    self.ws_import.append([None])
                self.ws_import.cell(row=import_row, column=5, value=serial)
            else:
                self.ws_import.cell(row=import_row, column=5, value=serial)
            item["serial"] = serial  # GUI zeigt Serial an

            # 2) Adresse setzen
            if self.ser.write_single(4, new_addr, unit=1).isError():
                raise RuntimeError("addr write")

            # 3) Buzzer ggf. deaktivieren
            if disable_bz:
                r = self.ser.read_holding(255, unit=1)
                if r.isError():
                    raise RuntimeError(r)
                self.ser.write_single(
                    255, r.registers[0] & ~(1 << 9), unit=1
                )

            # 4) Reboot
            if self.ser.write_single(17, 42330, unit=1).isError():
                raise RuntimeError("restart")

            # 5) feste Pause → Sensor bootet
            time.sleep(BOOT_WAIT_S)

            self.log(f"Sensor {row-1} OK (SN={serial})")
            return True, serial

        except Exception as e:
            self.log(f"Config-Fehler: {e}")
            return False, None
