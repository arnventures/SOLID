"""
excel_service.py
----------------
Lädt Sensordaten + exportiert LED/Sensor-Daten für GoLabel.
"""
from __future__ import annotations
import openpyxl, os
from typing import Tuple, List

# ----------------------------------------------------------------------
# Sensorliste laden (GUI)
# ----------------------------------------------------------------------
def load_sensor_data(path: str) -> Tuple[openpyxl.Workbook,
                                         openpyxl.worksheet.worksheet.Worksheet,
                                         List[tuple]]:
    wb = openpyxl.load_workbook(
        path,
        data_only=True,
        keep_vba=os.path.splitext(path)[1].lower() == ".xlsm"
    )
    ws = wb["GAS"]
    sensors = []
    for row in range(2, 62):
        model = ws[f"A{row}"].value
        location = ws[f"C{row}"].value
        addr   = ws[f"B{row}"].value
        buzzer = ws[f"D{row}"].value
        serial = ws[f"E{row}"].value
        if addr:
            sensors.append((row, model, location, addr, buzzer, serial))
    return wb, ws, sensors

# ----------------------------------------------------------------------
# LED-Daten kopieren (nur Slave-ID > 0)
# ----------------------------------------------------------------------
_LED_SHEETS = {
    "LIGHTIMOK", "LIGHT", "Light",
    "LEDIMOK", "LED", "Led", "Light 24V"
}

def copy_led_data(src: str, dst: str) -> bool:
    """
    Kopiert alle Zeilen mit Slave-ID > 0 aus dem LED-Blatt in die
    PowerAutomate-Vorlage.  Rückgabe True, wenn mind. eine Zeile kopiert wurde.
    """
    try:
        # Quelle
        swb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        src_sheet = next(
            (swb[name] for name in swb.sheetnames
             if name.strip() in _LED_SHEETS),
            None
        )
        if src_sheet is None:
            print("[LED-Copy] Kein LIGHT/LED-Blatt gefunden."); return False

        # Ziel vorbereiten
        if not os.path.exists(dst):
            openpyxl.Workbook().save(dst)        # leere Datei anlegen
        dwb = openpyxl.load_workbook(dst)
        d = dwb.active
        d.delete_rows(1, d.max_row)
        d.delete_cols(1, d.max_column)

        # Header kopieren
        d.append([c.value for c in src_sheet[1]])

        # Daten
        dest_row = 2
        for row_vals in src_sheet.iter_rows(min_row=2, values_only=True):
            sid = row_vals[1]
            try:
                sid_int = int(sid)
            except (TypeError, ValueError):
                continue
            if sid_int > 0:
                for col, val in enumerate(row_vals, 1):
                    if val is not None:
                        d.cell(dest_row, col).value = val
                dest_row += 1

        dwb.save(dst)
        return dest_row > 2               # True, wenn mind. 1 Zeile kopiert

    except Exception as e:
        print("[LED-Copy] Fehler:", e)
        return False

# ----------------------------------------------------------------------
# Sensor-Daten kopieren (unverändert)
# ----------------------------------------------------------------------
def copy_sensor_data(src: str, dst: str) -> bool:
    """
    Kopiert das GAS-Blatt 1:1 in die PowerAutomate-Sensorvorlage.
    """
    try:
        swb = openpyxl.load_workbook(src, read_only=True, data_only=True)
        s = swb["GAS"]
        dwb = openpyxl.load_workbook(dst)
        d = dwb.active
        d.delete_rows(1, d.max_row)
        d.delete_cols(1, d.max_column)

        for r, row in enumerate(s.iter_rows(min_row=2, values_only=True), 1):
            for c, val in enumerate(row, 1):
                if val is not None:
                    d.cell(r, c).value = val
        dwb.save(dst)
        return True
    except Exception as e:
        print("Sensor-Copy-Error:", e)
        return False
